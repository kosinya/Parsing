import time
import csv
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from openpyxl import load_workbook


# Сохранение в Excel
def fill_excel_template(csv_filename, excel_template_filename, output_filename):
    try:
        csv_data = []
        with open(csv_filename, 'r', encoding='utf-8-sig') as csv_file:
            csv_reader = csv.DictReader(csv_file)
            for row in csv_reader:
                csv_data.append(row)

        wb = load_workbook(excel_template_filename)

        sheet_name = 'Реестр средств размещения'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            print(f"Лист '{sheet_name}' не найден. Используется активный лист: '{ws.title}'")

        start_row = 3
        for i, row_data in enumerate(csv_data, start=start_row):
            ws.cell(row=i, column=1, value=row_data.get('Название гостиницы', ''))
            ws.cell(row=i, column=2, value=row_data.get('Адрес отеля', ''))
            ws.cell(row=i, column=3, value=row_data.get('Тип средства размещения', ''))
            ws.cell(row=i, column=4, value=row_data.get('Звёздность', ''))
            ws.cell(row=i, column=5, value=row_data.get('Статус', ''))
            ws.cell(row=i, column=6, value=row_data.get('Email', ''))
            ws.cell(row=i, column=7, value=row_data.get('Телефон отеля', ''))
            ws.cell(row=i, column=8, value=row_data.get('Телефон владельца', ''))
            ws.cell(row=i, column=9, value=row_data.get('ИНН', ''))
            ws.cell(row=i, column=10, value=row_data.get('Адрес владельца', ''))
            ws.cell(row=i, column=11, value=row_data.get('Номерной фонд', ''))
            ws.cell(row=i, column=12, value=row_data.get('ФИО руководителя', ''))

        wb.save(output_filename)
        print(f"Шаблон заполнен и сохранен как: {output_filename}")

    except FileNotFoundError as e:
        print(f"Ошибка: Файл не найден - {e}")
    except Exception as e:
        print(f"Ошибка при заполнении шаблона Excel: {e}")


# Нажатие на кнопку "Показать еще" через JavaScript
def click_show_more_button(driver):
    try:
        show_more_button = driver.find_element(
            By.XPATH,
            "//button[contains(text(), 'Показать ещё')]"
        )

        if show_more_button.is_displayed() and show_more_button.is_enabled():
            driver.execute_script("arguments[0].click();", show_more_button)
            return True

    except Exception as e:
        print(f"\tКнопка 'Показать ещё' не найдена.")

    return False


def get_new_hotels_count(driver, previous_count):
    current_hotels = driver.find_elements(By.TAG_NAME, "hotels-resort-card")
    current_count = len(current_hotels)

    new_count = current_count - previous_count

    if new_count > 0:
        print(f"\nЗагружено новых отелей: {new_count}")
        print(f"Теперь всего: {current_count}")
        return True

    print("\nНовых отелей не загрузилось.")
    return False


def extract_hotel_data(driver):
    data = {}

    # Название
    try:
        name_elem = driver.find_element(By.TAG_NAME, "h1")
        data['name'] = name_elem.text.strip()
    except NoSuchElementException:
        data['name'] = ''

    # Адрес
    try:
        address_elem = driver.find_element(By.XPATH, "//p[contains(text(), 'Адрес:')]/following-sibling::p")
        data['address'] = address_elem.text.strip()
    except:
        data['address'] = ''

    # Тип средства размещения
    try:
        type_elem = driver.find_element(By.XPATH,
                                        "//p[contains(text(), 'Тип средства размещения:')]/following-sibling::p")
        data['type'] = type_elem.text.strip()
    except:
        data['type'] = ''

    # Количество звезд
    try:
        stars = driver.find_elements(By.CSS_SELECTOR, "stars-bar ui-icon[icon='star'].active")
        data['stars'] = 5 - len(stars)
    except:
        data['stars'] = 0

    # Владелец
    try:
        owner_elem = driver.find_element(By.XPATH, "//p[contains(text(), 'Владелец:')]/following-sibling::p")
        data['owner'] = owner_elem.text.strip()
    except:
        data['owner'] = ''

    # Раскрытие модального окна с дополнительной информацией
    try:
        button = driver.find_element(By.XPATH, "//button[contains(., 'Информация')]")
        driver.execute_script("arguments[0].click();", button)
        print("\tКнопка 'Информация' нажата")
    except:
        print("\tКнопка не найдена, возможно страница не загрузилась")

    time.sleep(3)

    modal = driver.find_element(By.CSS_SELECTOR, "hotel-info-drawer")

    # ИНН
    try:
        inn_element = driver.find_element(
            By.XPATH,
            "//p[text()='ИНН']/ancestor::mat-expansion-panel//p[@class='content ng-star-inserted']"
        )
        inn_value = driver.execute_script("return arguments[0].textContent;", inn_element)
        data['inn'] = inn_value
    except:
        data['inn'] = ''

    # Адрес владельца
    try:
        owner_address_element = modal.find_element(By.XPATH,
                                                   ".//p[text()='Адрес владельца']/following::p[@class='content ng-star-inserted'][1]")
        owner_address_value = driver.execute_script("return arguments[0].textContent;", owner_address_element)
        data['owner_address'] = owner_address_value
    except:
        data['owner_address'] = ''

    # Телефон владельца
    try:
        owner_phone_element = modal.find_element(By.XPATH,
                                                 ".//p[text()='Телефон владельца']/following::p[@class='content ng-star-inserted'][1]")
        owner_address_value = driver.execute_script("return arguments[0].textContent;", owner_phone_element)
        data['owner_phone'] = owner_address_value
    except:
        data['owner_phone'] = ''

    # Телефон средства размещения
    try:
        hotel_phone_element = modal.find_element(By.XPATH, ".//p[text()='Телефон']/following::p[@class='content ng-star-inserted'][1]")
        hotel_phone_value = driver.execute_script("return arguments[0].textContent;", hotel_phone_element)
        data['hotel_phone'] = hotel_phone_value
    except:
        data['hotel_phone'] = ''

    # Почта средства размещения
    try:
        hotel_email_element = modal.find_element(By.XPATH,
                                                 ".//p[text()='Электронная почта']/following::p[@class='content ng-star-inserted'][1]")
        hotel_email_value = driver.execute_script("return arguments[0].textContent;", hotel_email_element)
        data['hotel_email'] = hotel_email_value
    except:
        data['hotel_email'] = ''

    button = driver.find_element(By.CSS_SELECTOR, "ui-icon[icon='characteristic']")
    driver.execute_script("arguments[0].closest('button').click();", button)

    time.sleep(1)

    try:
        button = driver.find_element(By.XPATH, "//button[contains(text(), 'Характеристики')]")
        driver.execute_script("arguments[0].click();", button)

        time.sleep(1)
    except Exception as e:
        print('Вторая кнопка "Характеристики" не найдена')

    # Номерной фонд
    try:
        element = driver.find_element(By.CSS_SELECTOR, "flex-block.room-counter")
        number = element.text.strip()
        data['num'] = number
    except:
        data['num'] = ''

    return data